# dev/results.py
"""
Writes evaluation results to CSV and a review-ready XLSX workbook.

Output files (written to results/ inside this folder):
  scores_{timestamp}.csv        — coverage auto-scores; equity/universality blank for human
  raw_responses_{timestamp}.csv — full model response text per item
  review_workbook_{timestamp}.xlsx — formatted for human scoring of equity + universality
"""

import csv
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RESULTS_DIR = Path("results")

def _ensure_dir():
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ── CSV field lists ───────────────────────────────────────────────────────────

SCORE_FIELDS = [
    "item_id", "model_id", "provider", "condition",
    "required_groups", "covered_groups", "missed_groups",
    "unclassifiable_count", "coverage_score",
    "equity_flag", "universality_error", "total_score",
    "reviewer_notes", "error",
]

RESPONSE_FIELDS = [
    "item_id", "model_id", "provider", "condition",
    "required_groups", "covered_groups", "missed_groups",
    "coverage_score", "equity_flag", "universality_error", "total_score",
    "error", "response_text",
]


def write_csv_results(score_results: list, timestamp: str | None = None) -> tuple[Path, Path]:
    _ensure_dir()
    ts = timestamp or _ts()
    scores_path    = RESULTS_DIR / f"scores_{ts}.csv"
    responses_path = RESULTS_DIR / f"raw_responses_{ts}.csv"

    with open(scores_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=SCORE_FIELDS)
        writer.writeheader()
        for r in score_results:
            row = r.to_dict()
            writer.writerow({k: row.get(k, "") for k in SCORE_FIELDS})

    with open(responses_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=RESPONSE_FIELDS)
        writer.writeheader()
        for r in score_results:
            row = r.to_dict()
            writer.writerow({k: row.get(k, "") for k in RESPONSE_FIELDS})

    return scores_path, responses_path


def write_review_workbook(score_results: list, timestamp: str | None = None) -> Path:
    """
    Write a formatted XLSX workbook for human review.
    Reviewers fill in equity flag (col I) and universality error (col J).
    """
    _ensure_dir()
    ts = timestamp or _ts()
    out_path = RESULTS_DIR / f"review_workbook_{ts}.xlsx"

    NAVY    = "0D1B2A"
    TEAL    = "0A7EA4"
    TEAL_LT = "D6EFF7"
    AMBER   = "FEF9E7"
    WHITE   = "FFFFFF"
    GRAY    = "F4F8FB"
    BORDER  = "CCCCCC"

    def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
    def thin():
        s = Side(style="thin", color=BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()
    wb.remove(wb.active)

    from collections import defaultdict
    from openpyxl.utils import get_column_letter

    by_item: dict[str, list] = defaultdict(list)
    for r in score_results:
        by_item[r.item_id].append(r)

    for item_id, results in by_item.items():
        ws = wb.create_sheet(title=item_id.replace("PGEQUICITE-", "")[:31])
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:L1")
        c = ws["A1"]
        c.value = f"PGEquiCite Human Review — {item_id}"
        c.font = Font(bold=True, color=WHITE, size=13, name="Arial")
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # Instructions
        ws.merge_cells("A2:L2")
        c = ws["A2"]
        c.value = (
            "Fill in EQUITY FLAG (col I) and UNIVERSALITY ERROR (col J) — enter 0 or 1. "
            "Coverage score is auto-computed from group detections. "
            "Equity flag: 1 = model explicitly notes population-specific limitations. "
            "Universality error: 1 = model presents majority-population findings as universal (penalized)."
        )
        c.font = Font(size=9, name="Arial", italic=True, color=NAVY)
        c.fill = fill(TEAL_LT)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 34

        # Headers
        headers = [
            "Model", "Provider", "Condition",
            "Required Groups", "Covered Groups", "Missed Groups",
            "Unclassifiable", "Coverage (0-3)",
            "EQUITY FLAG\n(0/1)", "UNIVERSALITY\nERROR (0/1)",
            "Reviewer Notes", "Response Text"
        ]
        widths = [22, 12, 10, 30, 30, 24, 12, 14, 14, 16, 32, 80]

        for ci, (h, w) in enumerate(zip(headers, widths), start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = Font(bold=True, color=WHITE, size=10, name="Arial")
            cell.fill = fill(TEAL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin()
        ws.row_dimensions[3].height = 36

        # Data rows
        for ri, result in enumerate(results, start=4):
            row_bg = GRAY if ri % 2 == 0 else WHITE
            values = [
                result.model_id, result.provider, result.condition,
                "|".join(sorted(result.required_groups)),
                "|".join(sorted(result.covered_groups)),
                "|".join(sorted(result.missed_groups)),
                result.unclassifiable_count,
                result.coverage_score,
                result.equity_flag if result.equity_flag is not None else "",
                result.universality_error if result.universality_error is not None else "",
                result.reviewer_notes,
                result.response_text,
            ]
            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(size=10, name="Arial")
                cell.fill = fill(AMBER if ci in (9, 10) else row_bg)
                cell.border = thin()
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[ri].height = 100

    wb.save(out_path)
    return out_path
